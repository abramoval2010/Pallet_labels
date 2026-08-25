# app/models/materials_db.py
import os
import sqlite3
import math
from contextlib import contextmanager
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class MaterialsDatabase:
    """Класс для работы с базой мастер-данных материалов"""

    def __init__(self, db_file):
        self.db_file = db_file
        self.init_database()
        print(f"Materials DB: {self.db_file}")

    @contextmanager
    def get_connection(self):
        """Получает соединение с базой данных"""
        conn = sqlite3.connect(self.db_file)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()

    def init_database(self):
        """Инициализирует базу данных и создает таблицу материалов"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS materials (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sap_code INTEGER UNIQUE NOT NULL,
                    material_name TEXT UNIQUE NOT NULL,
                    palletization INTEGER DEFAULT 0,
                    box_quantity INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            conn.commit()

            cursor.execute("SELECT COUNT(*) FROM materials")
            count = cursor.fetchone()[0]

            if count == 0:
                self._create_default_materials()

    def _create_default_materials(self):
        """Создает тестовые материалы"""
        default_materials = [
            {'sap_code': 10000001, 'material_name': 'Тестовый материал 1', 'palletization': 100, 'box_quantity': 10},
            {'sap_code': 10000002, 'material_name': 'Тестовый материал 2', 'palletization': 200, 'box_quantity': 20},
            {'sap_code': 10000003, 'material_name': 'Тестовый материал 3', 'palletization': 150, 'box_quantity': 15},
        ]
        for material in default_materials:
            self.add_material(material)

    def add_material(self, material_data):
        """Добавляет новый материал в базу"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO materials (sap_code, material_name, palletization, box_quantity)
                    VALUES (?, ?, ?, ?)
                ''', (
                    material_data['sap_code'],
                    material_data['material_name'],
                    material_data.get('palletization', 0),
                    material_data.get('box_quantity', 0)
                ))
                conn.commit()
                return True, "Материал успешно добавлен"
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint failed: materials.sap_code' in str(e):
                return False, f"Материал с SAP-кодом {material_data['sap_code']} уже существует"
            elif 'UNIQUE constraint failed: materials.material_name' in str(e):
                return False, f"Материал с названием '{material_data['material_name']}' уже существует"
            return False, f"Ошибка целостности данных: {str(e)}"
        except Exception as e:
            return False, f"Ошибка при добавлении материала: {str(e)}"

    def get_material_by_sap(self, sap_code):
        """Находит материал по SAP-коду"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM materials WHERE sap_code = ?", (sap_code,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def get_material_by_name(self, material_name):
        """Находит материал по названию"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM materials WHERE material_name = ?", (material_name,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def update_material(self, sap_code, new_data):
        """Обновляет данные материала"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()

                fields = []
                values = []

                for key in ['material_name', 'palletization', 'box_quantity']:
                    if key in new_data:
                        fields.append(f"{key} = ?")
                        values.append(new_data[key])

                values.append(sap_code)
                fields.append("updated_at = CURRENT_TIMESTAMP")

                query = f"UPDATE materials SET {', '.join(fields)} WHERE sap_code = ?"
                cursor.execute(query, values)
                conn.commit()
                return True, "Данные материала обновлены"
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint failed: materials.material_name' in str(e):
                return False, f"Материал с названием '{new_data.get('material_name', '')}' уже существует"
            return False, f"Ошибка целостности данных: {str(e)}"
        except Exception as e:
            return False, f"Ошибка при обновлении: {str(e)}"

    def delete_material(self, sap_code):
        """Удаляет материал"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM materials WHERE sap_code = ?", (sap_code,))
                conn.commit()
                return True, "Материал удален"
        except Exception as e:
            return False, f"Ошибка при удалении: {str(e)}"

    def get_all_materials(self):
        """Возвращает список всех материалов"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM materials ORDER BY sap_code")
            rows = cursor.fetchall()
            return [dict(row) for row in rows]

    def get_materials_count(self):
        """Возвращает количество материалов в базе"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM materials")
            return cursor.fetchone()[0]

    def validate_and_import_from_excel(self, filepath):
        """Проверяет и импортирует данные из Excel файла"""
        print(f"Importing Excel file: {filepath}")
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Определяем, есть ли заголовок
            is_header = False
            first_row = [cell.value for cell in ws[1]]
            if first_row:
                header_count = 0
                for cell in first_row[:4]:
                    if cell is not None and isinstance(cell, str) and not str(cell).isdigit():
                        header_count += 1
                if header_count >= 2:
                    is_header = True
                    start_row = 2
                else:
                    start_row = 1
            else:
                start_row = 1

            errors = []
            materials_to_add = []
            materials_to_update = []
            skipped_empty_box = 0
            has_box_quantity = ws.max_column >= 4

            for row_idx in range(start_row, ws.max_row + 1):
                sap_value = ws.cell(row=row_idx, column=1).value
                material_name = ws.cell(row=row_idx, column=2).value
                palletization_value = ws.cell(row=row_idx, column=3).value
                box_quantity_value = ws.cell(row=row_idx, column=4).value if has_box_quantity else None

                if sap_value is None:
                    errors.append(f"Row {row_idx}: SAP code cannot be empty")
                    continue

                try:
                    sap_code = int(sap_value)
                    if len(str(sap_code)) != 8:
                        errors.append(f"Row {row_idx}: SAP code must be 8 digits (current: {sap_code})")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: SAP code must be a number (current: {sap_value})")
                    continue

                material_name = str(material_name) if material_name else ''
                if not material_name.strip():
                    errors.append(f"Row {row_idx}: Material name cannot be empty")
                    continue
                material_name = material_name.strip()

                try:
                    palletization = int(palletization_value) if palletization_value is not None else 0
                    if palletization < 0:
                        errors.append(f"Row {row_idx}: Palletization cannot be negative (current: {palletization})")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Palletization must be a number (current: {palletization_value})")
                    continue

                box_quantity = None
                if has_box_quantity and box_quantity_value is not None:
                    try:
                        if isinstance(box_quantity_value, (int, float)):
                            box_quantity = int(box_quantity_value)
                        else:
                            box_quantity_str = str(box_quantity_value).strip().replace(',', '.')
                            if box_quantity_str:
                                box_quantity = int(float(box_quantity_str))

                        if box_quantity is not None and box_quantity < 0:
                            errors.append(f"Row {row_idx}: Box quantity cannot be negative (current: {box_quantity})")
                            box_quantity = None
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_idx}: Box quantity must be a number (current: {box_quantity_value})")
                        box_quantity = None
                elif has_box_quantity and box_quantity_value is None:
                    skipped_empty_box += 1

                existing = self.get_material_by_sap(sap_code)
                if existing:
                    new_data = {
                        'material_name': material_name,
                        'palletization': palletization
                    }
                    if box_quantity is not None:
                        new_data['box_quantity'] = box_quantity
                    materials_to_update.append({
                        'sap_code': sap_code,
                        'data': new_data
                    })
                else:
                    existing_name = self.get_material_by_name(material_name)
                    if existing_name:
                        errors.append(
                            f"Row {row_idx}: Material with name '{material_name}' already exists (SAP: {existing_name['sap_code']})")
                        continue

                    materials_to_add.append({
                        'sap_code': sap_code,
                        'material_name': material_name,
                        'palletization': palletization,
                        'box_quantity': box_quantity if box_quantity is not None else 0
                    })

            if errors:
                return False, "Errors in file:\n" + "\n".join(errors[:20]) + (
                    f"\n... and {len(errors) - 20} more errors" if len(errors) > 20 else "")

            updated_count = 0
            for material in materials_to_update:
                success, message = self.update_material(material['sap_code'], material['data'])
                if success:
                    updated_count += 1
                else:
                    errors.append(f"SAP {material['sap_code']}: {message}")

            added_count = 0
            for material in materials_to_add:
                success, message = self.add_material(material)
                if success:
                    added_count += 1
                else:
                    errors.append(f"SAP {material['sap_code']}: {message}")

            result_message = f"Processed: {added_count} added, {updated_count} updated"
            if skipped_empty_box > 0:
                result_message += f", {skipped_empty_box} box quantity updates skipped (empty cells)"
            if errors:
                result_message += f"\nErrors: " + "\n".join(errors[:5]) + (
                    f"\n... and {len(errors) - 5} more errors" if len(errors) > 5 else "")
                return False, result_message

            return True, result_message
        except Exception as e:
            print(f"Error importing: {e}")
            import traceback
            traceback.print_exc()
            return False, f"Error processing file: {str(e)}"

    def export_to_excel(self):
        """Экспортирует данные материалов в Excel файл"""
        try:
            materials = self.get_all_materials()
            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр сырья и материалов"

            headers = ['Номер SAP', 'Наименование', 'Палетизация', 'Вложение в короб']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_idx, material in enumerate(materials, 2):
                ws.cell(row=row_idx, column=1, value=material['sap_code'])
                ws.cell(row=row_idx, column=2, value=material['material_name'])
                ws.cell(row=row_idx, column=3, value=material['palletization'])
                ws.cell(row=row_idx, column=4, value=material['box_quantity'])

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20

            filename = f"Реестр_сырья_и_материалов_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.path.dirname(self.db_file), filename)
            wb.save(filepath)
            return filepath
        except Exception as e:
            print(f"Error exporting: {e}")
            return None